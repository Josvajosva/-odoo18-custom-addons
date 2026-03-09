from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import datetime, timedelta
import io
import tempfile
import binascii
import base64
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.debug('Cannot import openpyxl')
try:
    import csv
except ImportError:
    _logger.debug('Cannot import csv')
import openpyxl

class ImportCrmData(models.TransientModel):
    _name = "import.crm.data"
    _description = "CRM Data Import"

    file_select = fields.Binary(string="Select File", required=True)
    import_option = fields.Selection([
        ('xls', 'Excel File (XLS/XLSX)'),
        ('csv', 'CSV File')
    ], string='File Format', default='xls', required=True)

    def import_file(self):
        if not self.file_select:
            raise UserError(_("Please select a file to import."))

        _logger.info("Starting import process...")

        if self.import_option == 'csv':
            result = self.import_csv()
        elif self.import_option == 'xls':
            result = self.import_excel()
        else:
            raise UserError(_("Please select either CSV or Excel format!"))

        _logger.info("Import process completed")
        return result

    def import_csv(self):
        """Import data from CSV file"""
        try:
            csv_data = base64.b64decode(self.file_select)
            data_file = io.StringIO(csv_data.decode("utf-8"))
            data_file.seek(0)
            csv_reader = csv.reader(data_file, delimiter=',')
            file_reader = list(csv_reader)

        except Exception as e:
            raise UserError(_("Invalid CSV file! Error: %s") % str(e))

        if len(file_reader) < 2:
            raise UserError(_("CSV file is empty or has only headers."))

        headers = [header.strip().lower() for header in file_reader[0]]
        _logger.info("CSV Headers found: %s", headers)
        
        success_count = 0
        error_count = 0
        update_count = 0
        create_count = 0

        for i, row in enumerate(file_reader[1:], 1):
            if not any(row):
                continue
                
            if len(row) != len(headers):
                error_count += 1
                _logger.warning("Row %s has different number of columns than headers", i)
                continue

            try:
                values = dict(zip(headers, [str(field).strip() for field in row]))
                _logger.info("Processing row %s: %s", i, values)
                
                result = self.create_opportunity(values)
                if result:
                    if result.get('action') == 'update':
                        update_count += 1
                    else:
                        create_count += 1
                    success_count += 1
                else:
                    error_count += 1
                    
            except Exception as e:
                error_count += 1
                _logger.error("Error processing row %s: %s", i, str(e))

        return self.show_result(success_count, error_count, update_count, create_count)

    def import_excel(self):
        """Import data from Excel file"""
        try:
            file_data = base64.b64decode(self.file_select)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as fp:
                fp.write(file_data)
                file_path = fp.name
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Invalid Excel file! Error: %s") % str(e))

        if sheet.max_row < 2:
            raise UserError(_("Excel file is empty or has only headers."))

        headers = []
        success_count = 0
        error_count = 0

        for row_no, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_no == 1:
                headers = [str(cell).strip().lower() if cell is not None else '' for cell in row]
                _logger.info("Excel Headers found: %s", headers)
                continue

            row_values = [str(cell).strip() if cell is not None else '' for cell in row]
            if not any(row_values):
                continue
            if len(row_values) != len(headers):
                error_count += 1
                _logger.error("Row %s has different number of columns than headers", row_no)
                continue

            values = dict(zip(headers, row_values))

            try:
                result = self.create_opportunity(values)
                if result:
                    success_count += 1
                    _logger.info("Successfully processed row %s", row_no)
                else:
                    error_count += 1
                    _logger.error("Failed to process row %s", row_no)
            except UserError as ue:
                error_count += 1
                _logger.error("Row %s failed: %s", row_no, str(ue))
            except Exception as e:
                error_count += 1
                _logger.error("Unexpected error in row %s: %s", row_no, str(e))

        return self.show_result(success_count, error_count)

    def import_with_openpyxl(self, file_path):
        """Import data from Excel file using openpyxl"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            if sheet.max_row < 2:
                raise UserError(_("Excel file is empty or has only headers."))

            headers = []
            success_count = 0
            error_count = 0
            update_count = 0
            create_count = 0

            for row_no, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_no == 1:
                    headers = [str(cell).strip().lower() if cell is not None else '' for cell in row]
                    _logger.info("Excel Headers found: %s", headers)
                    continue

                try:
                    row_values = [str(cell).strip() if cell is not None else '' for cell in row]
                    
                    if not any(row_values):
                        continue
                        
                    if len(row_values) != len(headers):
                        error_count += 1
                        _logger.warning("Row %s has different number of columns than headers", row_no)
                        continue

                    values = dict(zip(headers, row_values))
                    _logger.info("Processing row %s: %s", row_no, values)
                    
                    result = self.create_opportunity(values)
                    if result:
                        if result.get('action') == 'update':
                            update_count += 1
                        else:
                            create_count += 1
                        success_count += 1
                    else:
                        error_count += 1
                        
                except Exception as e:
                    error_count += 1
                    _logger.error("Error processing row %s: %s", row_no, str(e))

            return self.show_result(success_count, error_count, update_count, create_count)
            
        except Exception as e:
            raise UserError(_("Error reading Excel file: %s") % str(e))

    def show_result(self, success_count, error_count, update_count=0, create_count=0):
        """Show import result to user"""
        message = _("Import completed!\n\nSuccessfully processed: %s opportunities\n- Created: %s\n- Updated: %s\nErrors: %s") % (
            success_count, create_count, update_count, error_count)
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Result'),
                'message': message,
                'sticky': True,
                'type': 'success' if success_count > 0 else 'warning',
            }
        }

    def get_stage_id(self, stage_name):
        """Find or create CRM stage based on name"""
        if not stage_name:
            return False
            
        stage = self.env['crm.stage'].search([
            ('name', '=ilike', stage_name.strip())
        ], limit=1)
        
        if stage:
            return stage.id
        else:
            try:
                new_stage = self.env['crm.stage'].create({
                    'name': stage_name.strip(),
                    'sequence': 100,
                })
                _logger.info("Created new stage: %s", stage_name)
                return new_stage.id
            except Exception as e:
                _logger.error("Error creating stage %s: %s", stage_name, str(e))
                return False

    def create_opportunity(self, values):
        """Create or Update Opportunity with all fields from your Excel file"""
        application_id = values.get('application id', '').strip() or values.get('application_id', '').strip()
        if not application_id:
            _logger.error("Application ID is required but not provided in the import data")
            return False
        
        existing_opportunity = self.env['crm.lead'].search([
            ('application_id', '=', application_id)
        ], limit=1)
        
        opportunity_name = values.get('opportunity', '').strip()
        if not opportunity_name:
            opportunity_name = values.get('customer', 'Imported Opportunity').strip()
        if not opportunity_name:
            opportunity_name = 'Imported Opportunity'
        
        partner_id = self.create_or_find_partner(values, application_id)
        
        default_team = self.env['crm.team'].search([('active', '=', True)], limit=1)
        if not default_team:
            default_team = self.env['crm.team'].create({'name': 'Sales', 'active': True})
        
        stage_name = values.get('stage', '').strip()
        stage_id = self.get_stage_id(stage_name)
        
        mi_type_mapping = {
            'drip irrigation': 'drip_irrigation',
            'sprinkler irrigation': 'sprinkler_irrigation',
            'rain gun': 'rain_gun',
            'micro sprinkler': 'micro_sprinkler',
            'bubbler irrigation': 'bubbler_irrigation',
            'drip to sprinkler system': 'drip_to_sprinkler',
        }
        
        irrigation_type = values.get('mi type', '').strip().lower()
        mi_type = mi_type_mapping.get(irrigation_type, False)
        
        caste_mapping = {
            'bc': 'bc',
            'mbc': 'mbc', 
            'sc': 'sc',
            'st': 'st',
            'oc': 'oc',
            'other': 'other',
        }
        caste_value = values.get('caste', '').strip().lower()
        caste = caste_mapping.get(caste_value, 'other')
        
        gender_mapping = {
            'male': 'male',
            'female': 'female',
            'other': 'other',
        }
        gender_value = values.get('gender', '').strip().lower()
        gender = gender_mapping.get(gender_value, 'other')
        
        opportunity_data = {
            'name': opportunity_name,
            'type': 'opportunity',
            'team_id': default_team.id,
            'user_id': self.env.uid,
            'father_name': values.get('father name', '').strip(),
            'caste': caste,
            'gender': gender,
            'district': values.get('district', '').strip(),
            'block': values.get('block', '').strip(),
            'village': values.get('village', '').strip(),
            'crop_name': values.get('crop', '').strip(),
            'total_area': self.parse_float(values.get('total area (ha)', '0')),
            'applied_area': self.parse_float(values.get('applied area (ha)', '0')),
            'department': values.get('department', '').strip(),
            'scheme': values.get('scheme', '').strip(),
            'mi_type': mi_type,
            'phone': values.get('mobile', '').strip(),
            #'mobile': values.get('mobile', '').strip(),
            'application_id': application_id,
            'work_order_no': values.get('work order no', '').strip() or values.get('work_order_no', '').strip() or values.get('work order number', '').strip() or values.get('work_order_number', '').strip(),
        }
        
        if stage_id:
            opportunity_data['stage_id'] = stage_id
        
        if partner_id:
            opportunity_data['partner_id'] = partner_id
        _logger.error("opportunity data: %s", opportunity_data)
        
        dealer_id = self.find_or_create_dealer(values.get('dealer', '').strip())
        _logger.error("opportunity dealer_id: %s", dealer_id)
        if dealer_id:
            opportunity_data['dealer_id'] = dealer_id
        _logger.error("opportunity data: %s", opportunity_data)
            
        _logger.error("opportunity existing_opportunity: %s", existing_opportunity)

        try:
            if existing_opportunity:
                _logger.info("Updating existing opportunity for Application ID: %s (Opportunity: %s)", application_id, existing_opportunity.name)
                existing_opportunity.write(opportunity_data)
                _logger.info("Successfully UPDATED opportunity: %s with Application ID: %s", opportunity_name, application_id)
                return {'action': 'update', 'record': existing_opportunity}
            else:
                _logger.error("CREATED new opportunity")
                opportunity = self.env['crm.lead'].create(opportunity_data)
                _logger.info("Successfully CREATED new opportunity: %s with Application ID: %s", opportunity_name, application_id)
                return {'action': 'create', 'record': opportunity}
        except Exception as e:
            _logger.error("Error processing opportunity %s: %s", opportunity_name, str(e))
            return False

    def parse_float(self, value):
        """Safely convert string to float"""
        try:
            return float(value) if value else 0.0
        except (ValueError, TypeError):
            return 0.0

    def create_or_find_partner(self, values, application_id):
        """Create or find customer as individual (person)"""
        customer_name = values.get('customer', '').strip()
        
        if not customer_name:
            return False
        
        partner_by_app_id = self.env['res.partner'].search([
            ('application_id', '=', application_id)
        ], limit=1)
        
        if partner_by_app_id:
            return partner_by_app_id.id
        
        partner = self.env['res.partner'].search([
            ('name', '=ilike', customer_name)
        ], limit=1)
        
        if partner:
            partner.write({'application_id': application_id})
            return partner.id
        else:
            caste_mapping = {
                'bc': 'bc', 'mbc': 'mbc', 'sc': 'sc', 
                'st': 'st', 'oc': 'oc', 'other': 'other',
            }
            gender_mapping = {
                'male': 'male', 'female': 'female', 'other': 'other',
            }
            
            caste_value = values.get('caste', '').strip().lower()
            gender_value = values.get('gender', '').strip().lower()
            
            partner_data = {
                'name': customer_name,
                'company_type': 'person',
                'father_name': values.get('father name', '').strip(),
                'caste': caste_mapping.get(caste_value, 'other'),
                'gender': gender_mapping.get(gender_value, 'other'),
                'street': f"{values.get('village', '')}, {values.get('block', '')}",
                'city': values.get('district', ''),
                'phone': values.get('mobile', '').strip(),
                #'mobile': values.get('mobile', '').strip(),
                'application_id': application_id,
            }
            
            try:
                new_partner = self.env['res.partner'].create(partner_data)
                _logger.info("Created new partner: %s with Application ID: %s", customer_name, application_id)
                return new_partner.id
            except Exception as e:
                _logger.error("Error creating partner %s: %s", customer_name, str(e))
                return False

    def find_or_create_dealer(self, dealer_name):
        """Find or create dealer partner as individual (person)"""
        if not dealer_name:
            return False
        
        dealer = self.env['res.partner'].search([
            ('name', '=ilike', dealer_name)
        ], limit=1)
        
        if dealer:
            if hasattr(self.env['res.partner'], 'is_dealer'):
                dealer.write({'is_dealer': True})
            return dealer.id
        else:
            dealer_data = {
                'name': dealer_name,
                'company_type': 'person',
            }
            
            if hasattr(self.env['res.partner'], 'is_dealer'):
                dealer_data['is_dealer'] = True
            
            try:
                new_dealer = self.env['res.partner'].create(dealer_data)
                _logger.info("Created new dealer (individual): %s", dealer_name)
                return new_dealer.id
            except Exception as e:
                _logger.error("Error creating dealer %s: %s", dealer_name, str(e))
                return False

class ImportInvoiceData(models.TransientModel):
    _name = "import.invoice.data"
    _description = "Invoice Data Import"

    file_select = fields.Binary(string="Select File", required=True)
    import_option = fields.Selection([
        ('xls', 'Excel File (XLS/XLSX)'),
        ('csv', 'CSV File')
    ], string='File Format', default='xls', required=True)
    
    def parse_float(self, value):
        """Safely convert string to float"""
        try:
            return float(value) if value else 0.0
        except (ValueError, TypeError):
            return 0.0
    
    def import_file(self):
        if not self.file_select:
            raise UserError(_("Please select a file to import."))

        _logger.info("Starting invoice import process...")

        if self.import_option == 'csv':
            result = self.import_csv()
        elif self.import_option == 'xls':
            result = self.import_excel()
        else:
            raise UserError(_("Please select either CSV or Excel format!"))

        _logger.info("Invoice import process completed")
        return result

    def import_csv(self):
        """Import data from CSV file (stop immediately on first error)"""
        try:
            csv_data = base64.b64decode(self.file_select)
            data_file = io.StringIO(csv_data.decode("utf-8"))
            data_file.seek(0)
            csv_reader = csv.reader(data_file, delimiter=',')
            file_reader = list(csv_reader)
            success_count = len(file_reader)-1
        except Exception as e:
            raise UserError(_("Invalid CSV file! Error: %s") % str(e))

        if len(file_reader) < 2:
            raise UserError(_("CSV file is empty or has only headers."))

        headers = [header.strip().lower() for header in file_reader[0]]
        _logger.info("CSV Headers found: %s", headers)

        for i, row in enumerate(file_reader[1:], 2):
            if not any(row):
                continue
            if len(row) != len(headers):
                raise UserError(_("Row %s has different number of columns than headers") % i)

            values = dict(zip(headers, [str(field).strip() for field in row]))

            try:
                self.create_invoice(values)
            except UserError as ue:
                raise UserError(_("Row %s failed: %s") % (i, str(ue)))
            except Exception as e:
                raise UserError(_("Unexpected error in row %s: %s") % (i, str(e)))

        return self.show_result(success_count, 0)


    def import_excel(self):
        """Import data from Excel file (stop immediately on first error)"""
        try:
            file_data = base64.b64decode(self.file_select)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as fp:
                fp.write(file_data)
                file_path = fp.name
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Invalid Excel file! Error: %s") % str(e))

        if sheet.max_row < 2:
            raise UserError(_("Excel file is empty or has only headers."))

        headers = []
        success_count = 0
        for row_no, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_no == 1:
                headers = [str(cell).strip().lower() if cell is not None else '' for cell in row]
                _logger.info("Excel Headers found: %s", headers)
                continue

            row_values = [str(cell).strip() if cell is not None else '' for cell in row]
            if not any(row_values):
                continue
            if len(row_values) != len(headers):
                raise UserError(_("Row %s has different number of columns than headers") % row_no)

            values = dict(zip(headers, row_values))

            try:
                self.create_invoice(values)
                success_count = row_no-1
            except UserError as ue:
                _logger.error("Row %s failed: %s", row_no, str(ue))
                raise ue
            except Exception as e:
                raise UserError(_("Unexpected error in row %s: %s") % (row_no, str(e)))

        return self.show_result(success_count, 0)

    def import_with_openpyxl(self, file_path):
        """Import data from Excel file using openpyxl"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            if sheet.max_row < 2:
                raise UserError(_("Excel file is empty or has only headers."))

            headers = []
            success_count = 0
            error_count = 0

            for row_no, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_no == 1:
                    headers = [str(cell).strip().lower() if cell is not None else '' for cell in row]
                    _logger.info("Excel Headers found: %s", headers)
                    continue

                try:
                    row_values = [str(cell).strip() if cell is not None else '' for cell in row]
                    
                    if not any(row_values):
                        continue
                        
                    if len(row_values) != len(headers):
                        error_count += 1
                        _logger.warning("Row %s has different number of columns than headers", row_no)
                        continue

                    values = dict(zip(headers, row_values))
                    _logger.info("Processing row %s: %s", row_no, values)
                    
                    record = self.create_invoice(values)
                    if record:
                        success_count += 1
                    else:
                        error_count += 1
                        
                except Exception as e:
                    error_count += 1
                    _logger.error("Error processing row %s: %s", row_no, str(e))

            return self.show_result(success_count, error_count)
            
        except Exception as e:
            raise UserError(_("Error reading Excel file: %s") % str(e))

    def show_result(self, success_count, error_count):
        """Show import result to user"""
        message = _("Import completed!\n\nSuccessfully created: %s invoices\nErrors: %s") % (success_count, error_count)
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Result'),
                'message': message,
                'sticky': True,
                'type': 'success' if success_count > 0 else 'warning',
            }
        }

    def parse_date(self, date_str):
        """Safely convert string to date"""
        if not date_str or date_str == 'None' or date_str.lower() == 'null' or date_str.strip() == '':
            return False
            
        try:
            date_str = str(date_str).split(' ')[0].strip()
            
            formats = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d']
            for fmt in formats:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            
            return False
        except Exception as e:
            _logger.warning("Error parsing date '%s': %s", date_str, str(e))
            return False

    # def get_work_status_id(self, status_name):
    #     """Map status name to work_status selection value"""
    #     status_mapping = {
    #         'final fund utr updated': 'fund_credited',
    #         'issued work order': 'issued_work_order',
    #         'work completed': 'work_completed',
    #         'work completion approved': 'work_completion_approved',
    #         'fund release recommended by block office': 'fund_release_recommended_block',
    #         'fund release recommended by district office': 'fund_release_recommended_district',
    #         'mi company rectification approved by block': 'rectification_approved_block',
    #         'mi company rectification reverted by block': 'rectification_reverted_block',
    #         'first fund release approved by state agri': 'fund_release_approved_agri',
    #         'first fund release approved by state horticulture': 'fund_release_approved_horti',
    #         'district first fund utr updated': 'utr_skipped',
    #         'joint verification completed': 'joint_verification_completed',
    #         'fund release proceeding completed': 'fund_release_proceeding_completed',
    #     }
        
    #     return status_mapping.get(status_name.strip().lower(), 'fund_credited')

    def get_work_status_id(self, status_name):
        """Map Excel/CSV status names to work_status technical keys"""
        if not status_name:
            return False

        mapping = {
            'application received': 'application_received',
            'approved by block officer': 'approved_by_block',
            'dd upload skipped': 'dd_upload_skipped',
            'district first fund credited (utr updated)': 'district_first_fund_credited',
            'district first fund proceeding completed': 'district_first_fund_proceeding_completed',
            'farmer acceptance letter uploaded': 'farmer_acceptance_uploaded',
            'final fund credited (utr updated)': 'fund_credited',
            'final fund release recommended by district office': 'final_fund_release_recommended_district',
            'first fund credited (utr updated)': 'first_fund_credited',
            'first fund proceeding completed': 'first_fund_proceeding_completed',
            'fund release proceeding completed': 'fund_release_proceeding_completed',
            'fund release recommended by block office': 'fund_release_recommended_block',
            'fund release recommended by district office': 'fund_release_recommended_district',
            'fund release verification by state agriculture': 'fund_release_approved_agri',
            'fund release verification by state horticulture': 'fund_release_approved_horti',
            'issued work order': 'issued_work_order',
            'joint verification completed': 'joint_verification_completed',
            'layout image and gps image uploaded': 'layout_image_uploaded',
            'mi company rectification approved by block': 'rectification_approved_block',
            'mi company rectification reverted by block': 'rectification_reverted_block',
            'pre inspection - request for revised quotation': 'preinspection_revised_quotation',
            'pre inspection approved': 'preinspection_approved',
            'quotation copy uploaded by mi company': 'quotation_copy_uploaded',
            'quotation prepared by mi company': 'quotation_prepared_mi',
            'reverted application rectified by block': 'reverted_rectified_by_block',
            'reverted application rectified by mi company': 'reverted_rectified_by_mi',
            'reverted by state agri / horti': 'reverted_state_agri_horti',
            'reverted by state agri / horti to block': 'reverted_state_agri_horti_block',
            'work completed': 'work_completed',
            'work completion approved': 'work_completion_approved',
        }

        key = status_name.strip().lower()
        return mapping.get(key, False)

    def get_system_type(self, system_type_value):
        """Map system type name to system_type selection value"""
        system_type_mapping = {
            'inline': 'inline',
            'inline system': 'inline',
            'online': 'online',
            'online system': 'online',
        }
        
        if not system_type_value:
            return 'inline'
            
        return system_type_mapping.get(system_type_value.strip().lower(), 'inline')

    def find_or_create_dealer(self, dealer_name):
        """Find or create dealer partner as individual (person)"""
        if not dealer_name:
            return False
        
        dealer = self.env['res.partner'].search([
            ('name', '=ilike', dealer_name)
        ], limit=1)
        
        if dealer:
            if hasattr(self.env['res.partner'], 'is_dealer'):
                dealer.write({'is_dealer': True})
            return dealer.id
        else:
            dealer_data = {
                'name': dealer_name,
                'company_type': 'person',
            }
            
            if hasattr(self.env['res.partner'], 'is_dealer'):
                dealer_data['is_dealer'] = True
            
            try:
                new_dealer = self.env['res.partner'].create(dealer_data)
                _logger.info("Created new dealer (individual): %s", dealer_name)
                return new_dealer.id
            except Exception as e:
                _logger.error("Error creating dealer %s: %s", dealer_name, str(e))
                return False

    def create_invoice(self, values):
        """Create or Update Invoice with fields that match your custom form"""
        application_id = values.get('application id', '').strip() or values.get('application_id', '').strip()
        
        if not application_id:
            _logger.error("Application ID is required but not provided in the import data")
            return False
        
        existing_invoice = self.env['account.move'].search([
            ('application_id', '=', application_id)
        ], limit=1)
        
        opportunity = self.env['crm.lead'].search([('application_id', '=', application_id)], limit=1)
        if not opportunity:
            raise UserError(_("No CRM record found for Application ID: %s.\nPlease create the CRM record before importing the invoice.") % application_id)

        if not opportunity.partner_id:
            raise UserError(_("CRM record for Application ID: %s has no linked customer.\nPlease link a customer to the CRM before importing the invoice.") % application_id)

        partner_id = opportunity.partner_id.id
        dealer_id = opportunity.dealer_id.id if opportunity.dealer_id and hasattr(self.env['account.move'], 'dealer_id') else False

        invoice_date = self.parse_date(values.get('invoice date', '') or values.get('invoice_date', ''))
        work_status_date = self.parse_date(values.get('current status date', '') or values.get('current_status_date', ''))
        
        fund_credit_date = self.parse_date(values.get('final fund utr date', '') or values.get('final_fund_utr_date', ''))
        
        supply_date = self.parse_date(values.get('supply date (eway bill)', '') or values.get('supply_date_eway_bill', '') or values.get('eway_date', ''))
        
        first_fund_utr_skipped_date = self.parse_date(values.get('first fund utr date', '') or values.get('first_fund_utr_skipped_date', ''))
        
        status_value = values.get('current status', '').strip() or values.get('current_status', '').strip()
        work_status = self.get_work_status_id(status_value)

        system_type_value = values.get('system type', '').strip() or values.get('system_type', '').strip()
        system_type = self.get_system_type(system_type_value)

        if system_type == 'online':
            product_name = "Online Drip system"
        else:
            product_name = "Inline emitting system"
        
        _logger.info("Looking for product: %s for system type: %s", product_name, system_type)


        product = self.env['product.product'].search([
            ('name', '=ilike', product_name)
        ], limit=1)
        
        if not product:
            product = self.env['product.product'].search([
                ('name', 'ilike', f'%{product_name}%')
            ], limit=1)
        
        if not product:
            _logger.error("Product '%s' not found in database for Application ID: %s", product_name, application_id)
            _logger.info("Available products: %s", self.env['product.product'].search([]).mapped('name'))
            return False

        _logger.info("Found product: %s (ID: %s)", product.name, product.id)

        invoice_value = self.parse_float(values.get('invoice value', '0') or values.get('price_subtotal', '0') or values.get('amount', '0') or values.get('total', '0'))
        
        # invoice_data = {
        #     'move_type': 'out_invoice',
        #     'partner_id': partner_id,
        #     'invoice_date': invoice_date or fields.Date.today(),
        #     'date': invoice_date or fields.Date.today(),
        #     'application_id': application_id,
        #     'work_status': work_status,
        #     'work_date': work_status_date or invoice_date or fields.Date.today(),
        # }

        invoice_data = {
            'move_type': 'out_invoice',
            'partner_id': partner_id,
            'application_id': application_id,
            'work_status': work_status,
            'work_date': work_status_date or invoice_date or fields.Date.today(),
            'invoice_date': supply_date or invoice_date or fields.Date.today(),
            'date': supply_date or invoice_date or fields.Date.today(),
        }

        if hasattr(self.env['account.move'], 'system_type'):
            invoice_data['system_type'] = system_type
            _logger.info("Set system_type: %s for Application ID: %s", system_type, application_id)

        if dealer_id and hasattr(self.env['account.move'], 'dealer_id'):
            invoice_data['dealer_id'] = dealer_id
            _logger.info("Assigned dealer from CRM to invoice for Application ID: %s", application_id)

        if first_fund_utr_skipped_date and hasattr(self.env['account.move'], 'first_fund_utr_skipped_date'):
            invoice_data['first_fund_utr_skipped_date'] = first_fund_utr_skipped_date
            _logger.info("Set first_fund_utr_skipped_date: %s for Application ID: %s", first_fund_utr_skipped_date, application_id)

        current_status_remarks = values.get('current status remarks', '').strip() or values.get('current_status_remarks', '').strip()
        if current_status_remarks:
            invoice_data['narration'] = current_status_remarks

        if fund_credit_date and hasattr(self.env['account.move'], 'fund_credit_date'):
            invoice_data['fund_credit_date'] = fund_credit_date

        if supply_date and hasattr(self.env['account.move'], 'eway_date'):
            invoice_data['eway_date'] = supply_date

        if hasattr(self.env['account.move'], 'work_completion_approved_date'):
            work_completion_approved_date = self.parse_date(values.get('work completion approved date', '') or values.get('work_completion_approved_date', ''))
            invoice_data['work_completion_approved_date'] = work_completion_approved_date or invoice_date or fields.Date.today()

        eway_number = values.get('eway bill number', '').strip() or values.get('eway_number', '').strip()
        if eway_number and hasattr(self.env['account.move'], 'eway_number'):
            invoice_data['eway_number'] = eway_number

        date_fields_mapping = {
            'application_received_date': ['application received date', 'application_received_date'],
            'quotation_date': ['quotation date', 'quotation_date'],
            'work_order_date': ['work order date', 'work_order_date'],
            'work_order_completion_date': ['work order completion date', 'work_order_completion_date'],
            'issued_work_order_date': ['issued work order date', 'issued_work_order_date'],
            'first_fund_utr_skipped_date': ['district first fund utr skipped date', 'first_fund_utr_skipped_date'],
            'joint_verification_completed_date': ['jv recommended date', 'jv_recommended_date'],
            'eway_date': ['supply date (eway bill)', 'supply_date_eway_bill', 'eway_date'],
        }

        for field_name, possible_keys in date_fields_mapping.items():
            if hasattr(self.env['account.move'], field_name):
                for key in possible_keys:
                    if key in values and values[key]:
                        date_value = self.parse_date(values[key])
                        if date_value:
                            invoice_data[field_name] = date_value
                            break

        text_fields_mapping = {
            'department': ['department', 'department'],
            'scheme': ['scheme', 'scheme'],
            'survey_number': ['survey number', 'survey_number'],
            'sub_division_no': ['sub division no', 'sub_division_no'],
            'crop_name': ['crop name', 'crop_name'],
            'block': ['block', 'block'],
            'eway_number': ['eway bill number', 'eway_number'],
        }

        for field_name, possible_keys in text_fields_mapping.items():
            if hasattr(self.env['account.move'], field_name):
                for key in possible_keys:
                    if key in values and values[key]:
                        invoice_data[field_name] = values[key].strip()
                        break

        try:
            if existing_invoice:
                _logger.info("Updating existing invoice for Application ID: %s (Invoice: %s)", application_id, existing_invoice.name)
                
                if existing_invoice.state in ['posted', 'cancel']:
                    existing_invoice.button_draft()
                #raise UserError(_("Invalid data: %s") % str(invoice_data))
                existing_invoice.write(invoice_data)

                if existing_invoice.invoice_line_ids:
                    invoice_line = existing_invoice.invoice_line_ids[0]
                    invoice_line.write({
                        'product_id': product.id,
                        'name': product.name,
                        'quantity': 1,
                        'price_unit': invoice_value,
                        'tax_ids': [(5, 0, 0)],
                    })
                else:
                    existing_invoice.write({
                        'invoice_line_ids': [(0, 0, {
                            'product_id': product.id,
                            'name': product.name,
                            'quantity': 1,
                            'price_unit': invoice_value,
                            'tax_ids': [(5, 0, 0)],
                        })]
                    })
                
                if hasattr(existing_invoice, '_compute_commission_deduction'):
                    existing_invoice._compute_commission_deduction()
                
                #if existing_invoice.state == 'draft':
                #    existing_invoice.action_post()
                    
                _logger.info("Successfully UPDATED invoice for Application ID: %s with invoice value: %s", application_id, invoice_value)
                return existing_invoice
            else:
                invoice_data['invoice_line_ids'] = [(0, 0, {
                    'product_id': product.id,
                    'name': product.name,
                    'quantity': 1,
                    'price_unit': invoice_value,
                    'tax_ids': [(5, 0, 0)],
                })]
                #raise UserError(_("Invalid data create: %s") % str(invoice_data))
                invoice = self.env['account.move'].create(invoice_data)

                #if invoice.state == 'draft':
                #    invoice.action_post()
                    
                _logger.info("Successfully CREATED new invoice for Application ID: %s with invoice value: %s", application_id, invoice_value)
                return invoice
                
        except Exception as e:
            _logger.error("Error processing invoice for Application ID %s: %s", application_id, str(e))
            return False

    def find_partner_by_application_id(self, application_id):
        """Find partner by application ID"""
        if not application_id:
            return False
            
        partner = self.env['res.partner'].search([
            ('application_id', '=', application_id.strip())
        ], limit=1)
        
        return partner.id if partner else False

    def find_partner_from_crm(self, application_id):
        """Find partner from CRM opportunities by application ID"""
        if not application_id:
            return False
            
        opportunity = self.env['crm.lead'].search([
            ('application_id', '=', application_id.strip())
        ], limit=1)
        
        if opportunity and opportunity.partner_id:
            return opportunity.partner_id.id
            
        return False

    def create_partner_for_invoice(self, application_id, values):
        """Create a new partner for the invoice"""
        try:
            partner_data = {
                'name': f"Customer-{application_id}",
                'application_id': application_id,
                'company_type': 'person',
                'customer_rank': 1,
            }
            partner = self.env['res.partner'].create(partner_data)
            _logger.info("Created new partner for Application ID: %s", application_id)
            return partner.id
        except Exception as e:
            _logger.error("Error creating partner for Application ID %s: %s", application_id, str(e))
            return False
